import openpyxl

# Function to load student data from the Excel file
def load_student_data(excel_file):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook["Sheet1"]

    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    return data

# Function to calculate points and update the leaderboard
def update_leaderboard(student_data):
    headers = student_data[0]
    student_data = student_data[1:]

    # Calculate points based on the average score for each student
    points_column = ["Points"]
    for student in student_data:
        scores = student[1:]
        average_score = sum(scores) / len(scores)
        points = sum([score - average_score for score in scores])
        student.append(points)

    # Sort students by points in descending order
    student_data.sort(key=lambda x: x[-1], reverse=True)

    return [headers] + student_data

def save_leaderboard(excel_file, updated_leaderboard):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet2"

    for row in updated_leaderboard:
        sheet.append(row)

    workbook.save(excel_file)

def main():
    excel_file = "your_excel_file.xlsx"  # Replace with your Excel file path

    # Load student data from Sheet1
    student_data = load_student_data(excel_file)

    # Update the leaderboard
    updated_leaderboard = update_leaderboard(student_data)

    # Save the updated leaderboard to Sheet2
    save_leaderboard(excel_file, updated_leaderboard)

if __name__ == "__main__":
    main()
